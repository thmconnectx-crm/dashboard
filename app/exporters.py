from datetime import date, datetime
from io import BytesIO
import math

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_RIGHT
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfgen import canvas as pdf_canvas
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from app.models import CampaignMetric


BRAND_DARK = colors.HexColor("#0B1117")
BRAND_PANEL = colors.HexColor("#111A22")
BRAND_GREEN = colors.HexColor("#35D092")
BRAND_CYAN = colors.HexColor("#39B8D8")
BRAND_AMBER = colors.HexColor("#F0B76B")
BRAND_LINE = colors.HexColor("#D9E1E8")
BRAND_SOFT = colors.HexColor("#F4F7FA")
BRAND_MUTED = colors.HexColor("#5B6874")
PDF_FONT_REGULAR = "Helvetica"
PDF_FONT_BOLD = "Helvetica-Bold"


def rows_to_dataframe(rows: list[CampaignMetric]) -> pd.DataFrame:
    data = [
        {
            "Data": row.date,
            "Plataforma": _platform_label(row.platform),
            "Conta": row.account_id,
            "Campanha": row.campaign_name,
            "Modelo da Campanha": _objective_label(getattr(row, "campaign_objective", "")),
            "Objetivo técnico": getattr(row, "campaign_objective", ""),
            "Alcance": getattr(row, "reach", 0),
            "Impressões": row.impressions,
            "Frequência": _safe_div(row.impressions, getattr(row, "reach", 0)),
            "Cliques": row.clicks,
            "CTR (%)": row.ctr,
            "CPC": row.cpc,
            "Investimento": row.spend,
            "Mensagens": getattr(row, "messages", 0.0),
            "Custo por Mensagem": getattr(row, "cost_per_message", 0.0),
            "Conversões": row.conversions,
            "Custo/Conv.": row.cost_per_conversion,
            "Valor de Conversão": row.conversion_value,
            "ROAS": row.roas,
        }
        for row in rows
    ]
    return pd.DataFrame(data)


def build_excel(rows: list[CampaignMetric]) -> bytes:
    df = rows_to_dataframe(rows)
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        if df.empty:
            pd.DataFrame([{"Status": "Nenhum dado encontrado para os filtros selecionados."}]).to_excel(
                writer, index=False, sheet_name="Resumo"
            )
        else:
            platform_summary = _aggregate(df, ["Plataforma"])
            campaign_summary = _aggregate(df, ["Plataforma", "Conta", "Modelo da Campanha", "Campanha"]).sort_values(
                by="Investimento", ascending=False
            )
            daily_summary = _aggregate(df, ["Data", "Plataforma"]).sort_values(by=["Data", "Plataforma"])

            overview = pd.DataFrame(
                [
                    {"Indicador": "Investimento total", "Valor": _money(float(df["Investimento"].sum()))},
                    {"Indicador": "Conversas iniciadas", "Valor": _number(float(df["Mensagens"].sum()))},
                    {"Indicador": "Custo por conversa", "Valor": _money(_safe_div(df["Investimento"].sum(), df["Mensagens"].sum()))},
                    {"Indicador": "Alcance", "Valor": _integer(float(df["Alcance"].sum()))},
                    {"Indicador": "Impressões", "Valor": _integer(float(df["Impressões"].sum()))},
                    {"Indicador": "Frequência", "Valor": _ratio(_safe_div(df["Impressões"].sum(), df["Alcance"].sum()))},
                    {"Indicador": "CTR médio", "Valor": _percent(_safe_div(df["Cliques"].sum(), df["Impressões"].sum()) * 100)},
                    {"Indicador": "CPC médio", "Valor": _money(_safe_div(df["Investimento"].sum(), df["Cliques"].sum()))},
                ]
            )
            overview.to_excel(writer, index=False, sheet_name="Resumo")
            platform_summary.to_excel(writer, index=False, sheet_name="Plataformas")
            campaign_summary.to_excel(writer, index=False, sheet_name="Campanhas")
            daily_summary.to_excel(writer, index=False, sheet_name="Evolução diária")
            df.sort_values(by=["Data", "Plataforma", "Campanha"]).to_excel(writer, index=False, sheet_name="Histórico")

        for worksheet in writer.book.worksheets:
            _style_worksheet(worksheet)
    return output.getvalue()


def build_pdf(
    rows: list[CampaignMetric],
    start_date: date | None = None,
    end_date: date | None = None,
    previous_rows: list[CampaignMetric] | None = None,
) -> bytes:
    _register_canvas_fonts()
    df = rows_to_dataframe(rows)
    period = _period_label(start_date, end_date)
    if df.empty:
        return _build_empty_premium_pdf(period)

    previous_df = rows_to_dataframe(previous_rows or [])
    campaign_summary = _aggregate(df, ["Plataforma", "Conta", "Modelo da Campanha", "Campanha"]).sort_values(by="Investimento", ascending=False)
    daily_summary = _aggregate(df, ["Data", "Plataforma"]).sort_values(by=["Data", "Plataforma"])
    return _build_premium_pdf_canvas(df, previous_df, campaign_summary, daily_summary, period)


def _register_canvas_fonts() -> None:
    global PDF_FONT_REGULAR, PDF_FONT_BOLD
    if PDF_FONT_REGULAR != "Helvetica":
        return
    font_pairs = [
        (
            "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
            "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf",
        ),
        (
            "/usr/share/fonts/truetype/liberation2/LiberationSans-Regular.ttf",
            "/usr/share/fonts/truetype/liberation2/LiberationSans-Bold.ttf",
        ),
        (
            "C:/Windows/Fonts/segoeui.ttf",
            "C:/Windows/Fonts/segoeuib.ttf",
        ),
    ]
    for regular_path, bold_path in font_pairs:
        try:
            with open(regular_path, "rb"), open(bold_path, "rb"):
                pdfmetrics.registerFont(TTFont("ReportSans", regular_path))
                pdfmetrics.registerFont(TTFont("ReportSans-Bold", bold_path))
                PDF_FONT_REGULAR = "ReportSans"
                PDF_FONT_BOLD = "ReportSans-Bold"
                return
        except OSError:
            continue



def _build_premium_pdf_canvas(
    df: pd.DataFrame,
    previous_df: pd.DataFrame,
    campaign_summary: pd.DataFrame,
    daily_summary: pd.DataFrame,
    period: str,
) -> bytes:
    output = BytesIO()
    page_width, page_height = landscape(A4)
    canvas = pdf_canvas.Canvas(output, pagesize=landscape(A4))
    canvas.setTitle("Relatório premium de tráfego pago")
    canvas.setAuthor("Relatório executivo de mídia paga")

    totals = _totals(df)
    report_mode = _report_mode(campaign_summary)
    display_summary = _display_campaign_summary(campaign_summary, report_mode)
    focus_df = _focus_dataframe(df, report_mode)
    focus_totals = _totals(focus_df) if not focus_df.empty else totals
    previous = _totals(previous_df) if not previous_df.empty else {}
    previous_focus_df = _focus_dataframe(previous_df, report_mode) if not previous_df.empty else previous_df
    previous_focus = _totals(previous_focus_df) if not previous_focus_df.empty else previous
    focus_daily_summary = _aggregate(focus_df, ["Data", "Plataforma"]).sort_values(by=["Data", "Plataforma"]) if not focus_df.empty else daily_summary
    margin = 22
    content_width = page_width - (margin * 2)

    _draw_page_background(canvas, page_width, page_height)
    _draw_canvas_header(canvas, margin, page_height - 92, content_width, 70, period, report_mode)

    card_y = page_height - 180
    card_gap = 10
    card_width = (content_width - card_gap * 3) / 4
    cards = _report_cards(report_mode, totals, focus_totals, previous, previous_focus)
    for index, (label, value, delta, previous_label) in enumerate(cards):
        _draw_metric_card(canvas, margin + index * (card_width + card_gap), card_y, card_width, 68, label, value, delta, previous_label)

    overview_y = page_height - 266
    _draw_campaign_overview(canvas, margin, overview_y, content_width, 76, display_summary, totals, focus_totals, report_mode)

    middle_y = page_height - 406
    left_width = content_width * 0.47
    right_x = margin + left_width + 14
    right_width = content_width - left_width - 14
    _draw_actions_panel(canvas, margin, middle_y, left_width, 122, totals, focus_totals, report_mode)
    _draw_daily_panel(canvas, right_x, middle_y, right_width, 122, focus_daily_summary, report_mode)

    _draw_featured_table(canvas, margin, 48, content_width, 142, display_summary.head(8), report_mode)
    _draw_canvas_footer(canvas, margin, page_width, period)

    canvas.showPage()
    canvas.save()
    return output.getvalue()


def _build_empty_premium_pdf(period: str) -> bytes:
    output = BytesIO()
    page_width, page_height = landscape(A4)
    canvas = pdf_canvas.Canvas(output, pagesize=landscape(A4))
    canvas.setTitle("Relatório premium de tráfego pago")
    margin = 22
    content_width = page_width - (margin * 2)
    _draw_page_background(canvas, page_width, page_height)
    _draw_canvas_header(canvas, margin, page_height - 92, content_width, 70, period, "messages")
    canvas.setFillColor(colors.white)
    canvas.roundRect(margin, page_height - 240, content_width, 110, 8, fill=1, stroke=0)
    canvas.setFillColor(BRAND_DARK)
    canvas.setFont(PDF_FONT_BOLD, 16)
    canvas.drawString(margin + 22, page_height - 175, "Nenhum dado encontrado para os filtros selecionados")
    canvas.setFillColor(BRAND_MUTED)
    canvas.setFont(PDF_FONT_REGULAR, 9)
    canvas.drawString(margin + 22, page_height - 196, "Verifique o período, a conta selecionada ou execute uma nova sincronização antes de exportar o relatório.")
    _draw_canvas_footer(canvas, margin, page_width, period)
    canvas.showPage()
    canvas.save()
    return output.getvalue()


def _report_mode(campaign_summary: pd.DataFrame) -> str:
    if campaign_summary.empty:
        return "messages"
    awareness_rows = campaign_summary[
        campaign_summary.apply(lambda row: _is_awareness_campaign(str(row.get("Modelo da Campanha", "")), str(row.get("Campanha", ""))), axis=1)
    ]
    return "awareness" if not awareness_rows.empty else "messages"


def _display_campaign_summary(campaign_summary: pd.DataFrame, report_mode: str) -> pd.DataFrame:
    if campaign_summary.empty:
        return campaign_summary
    if report_mode == "awareness":
        filtered = campaign_summary[
            campaign_summary.apply(lambda row: _is_awareness_campaign(str(row.get("Modelo da Campanha", "")), str(row.get("Campanha", ""))), axis=1)
        ].copy()
        filtered = filtered[~filtered["Campanha"].astype(str).map(_is_report_support_campaign)]
        return filtered if not filtered.empty else campaign_summary.head(1)
    filtered = campaign_summary[~campaign_summary["Campanha"].astype(str).map(_is_report_support_campaign)].copy()
    return filtered if not filtered.empty else campaign_summary


def _focus_dataframe(df: pd.DataFrame, report_mode: str) -> pd.DataFrame:
    if df.empty:
        return df
    if report_mode == "awareness":
        filtered = df[df.apply(lambda row: _is_awareness_campaign(str(row.get("Modelo da Campanha", "")), str(row.get("Campanha", ""))), axis=1)].copy()
        filtered = filtered[~filtered["Campanha"].astype(str).map(_is_report_support_campaign)]
        return filtered if not filtered.empty else df
    return df[~df["Campanha"].astype(str).map(_is_report_support_campaign)].copy()


def _is_awareness_campaign(model: str, campaign_name: str) -> bool:
    model_text = _normalize_text(model)
    name_text = _normalize_text(campaign_name)
    return (
        "reconhecimento" in model_text
        or "alcance" in model_text
        or "[rec]" in name_text
        or ".rec" in name_text
        or " rec " in f" {name_text} "
    )


def _is_report_support_campaign(campaign_name: str) -> bool:
    name_text = _normalize_text(campaign_name)
    return "copia" in name_text or "copy" in name_text


def _normalize_text(value: str) -> str:
    replacements = {
        "á": "a",
        "à": "a",
        "â": "a",
        "ã": "a",
        "é": "e",
        "ê": "e",
        "í": "i",
        "ó": "o",
        "ô": "o",
        "õ": "o",
        "ú": "u",
        "ç": "c",
    }
    text = str(value or "").strip().lower()
    for source, target in replacements.items():
        text = text.replace(source, target)
    return text


def _report_cards(
    report_mode: str,
    account_totals: dict[str, float],
    focus_totals: dict[str, float],
    previous: dict[str, float],
    previous_focus: dict[str, float],
) -> list[tuple[str, str, str, str]]:
    if report_mode == "awareness":
        frequency = _safe_div(focus_totals["impressions"], focus_totals["reach"])
        previous_frequency = _safe_div(previous_focus.get("impressions", 0), previous_focus.get("reach", 0))
        return [
            ("Valor investido", _money(account_totals["spend"]), _delta(account_totals["spend"], previous.get("spend", 0)), _previous_label(previous.get("spend", 0), _money)),
            ("Alcance", _integer(focus_totals["reach"]), _delta(focus_totals["reach"], previous_focus.get("reach", 0)), _previous_label(previous_focus.get("reach", 0), _integer)),
            ("Impressões", _integer(focus_totals["impressions"]), _delta(focus_totals["impressions"], previous_focus.get("impressions", 0)), _previous_label(previous_focus.get("impressions", 0), _integer)),
            ("Frequência", _ratio(frequency), _delta(frequency, previous_frequency), _previous_label(previous_frequency, _ratio)),
        ]
    return [
        ("Valor investido", _money(account_totals["spend"]), _delta(account_totals["spend"], previous.get("spend", 0)), _previous_label(previous.get("spend", 0), _money)),
        ("Conversas iniciadas", _number(focus_totals["messages"]), _delta(focus_totals["messages"], previous.get("messages", 0)), _previous_label(previous.get("messages", 0), _number)),
        ("Cliques", _integer(focus_totals["clicks"]), _delta(focus_totals["clicks"], previous.get("clicks", 0)), _previous_label(previous.get("clicks", 0), _integer)),
        ("Impressões", _integer(focus_totals["impressions"]), _delta(focus_totals["impressions"], previous.get("impressions", 0)), _previous_label(previous.get("impressions", 0), _integer)),
    ]


def _draw_page_background(canvas, page_width: float, page_height: float) -> None:
    canvas.setFillColor(colors.HexColor("#F3F6F8"))
    canvas.rect(0, 0, page_width, page_height, fill=1, stroke=0)
    canvas.setFillColor(colors.HexColor("#E7EDF2"))
    canvas.rect(0, 0, page_width, 26, fill=1, stroke=0)


def _draw_canvas_header(canvas, x: float, y: float, width: float, height: float, period: str, report_mode: str) -> None:
    canvas.setFillColor(BRAND_DARK)
    canvas.roundRect(x, y, width, height, 10, fill=1, stroke=0)
    canvas.setFillColor(BRAND_GREEN)
    canvas.rect(x, y + height - 4, width, 4, fill=1, stroke=0)
    canvas.setFillColor(colors.white)
    _draw_fitted_text(
        canvas,
        "Análise de desempenho de campanhas",
        x + 24,
        y + height - 31,
        width - 232,
        PDF_FONT_BOLD,
        19,
    )
    canvas.setFillColor(colors.HexColor("#C9D4DD"))
    _draw_fitted_text(
        canvas,
        "Meta Ads | Alcance, impressões, frequência, CPM, cliques e investimento"
        if report_mode == "awareness"
        else "Meta Ads | Campanhas de engajamento, conversas iniciadas, cliques, alcance e investimento",
        x + 24,
        y + 19,
        width - 232,
        PDF_FONT_REGULAR,
        9.2,
    )
    canvas.setFillColor(colors.HexColor("#19242E"))
    canvas.roundRect(x + width - 178, y + 16, 154, 38, 8, fill=1, stroke=0)
    canvas.setFillColor(colors.HexColor("#C9D4DD"))
    canvas.setFont(PDF_FONT_REGULAR, 8)
    canvas.drawRightString(x + width - 36, y + 39, "Período analisado")
    canvas.setFillColor(colors.white)
    canvas.setFont(PDF_FONT_BOLD, 9.4)
    canvas.drawRightString(x + width - 36, y + 24, period)


def _draw_metric_card(canvas, x: float, y: float, width: float, height: float, label: str, value: str, delta: str, previous: str) -> None:
    canvas.setFillColor(colors.white)
    canvas.roundRect(x, y, width, height, 8, fill=1, stroke=0)
    canvas.setStrokeColor(colors.HexColor("#DDE5EC"))
    canvas.roundRect(x, y, width, height, 8, fill=0, stroke=1)
    canvas.setFillColor(BRAND_MUTED)
    _draw_fitted_text(canvas, label.upper(), x + 12, y + height - 17, width - 24, PDF_FONT_BOLD, 8.3)
    canvas.setFillColor(BRAND_DARK)
    _draw_fitted_text(canvas, value, x + 12, y + height - 42, width - 24, PDF_FONT_BOLD, 17.2)
    if delta == "Novo período":
        canvas.setFillColor(BRAND_MUTED)
        _draw_fitted_text(canvas, "Sem comparação anterior", x + 12, y + 12, width - 24, PDF_FONT_REGULAR, 7.4)
        return
    canvas.setFillColor(BRAND_GREEN if not delta.startswith("-") else colors.HexColor("#D65F5F"))
    _draw_fitted_text(canvas, delta, x + 12, y + 19, width - 24, PDF_FONT_BOLD, 8)
    canvas.setFillColor(BRAND_MUTED)
    _draw_fitted_text(canvas, previous, x + 12, y + 8, width - 24, PDF_FONT_REGULAR, 7)


def _draw_campaign_overview(
    canvas,
    x: float,
    y: float,
    width: float,
    height: float,
    campaign_summary: pd.DataFrame,
    account_totals: dict[str, float],
    focus_totals: dict[str, float],
    report_mode: str,
) -> None:
    canvas.setFillColor(colors.white)
    canvas.roundRect(x, y, width, height, 8, fill=1, stroke=0)
    canvas.setStrokeColor(colors.HexColor("#DDE5EC"))
    canvas.roundRect(x, y, width, height, 8, fill=0, stroke=1)
    canvas.setFillColor(BRAND_DARK)
    canvas.setFont(PDF_FONT_BOLD, 11.5)
    canvas.drawString(x + 12, y + height - 18, "Campanhas")
    canvas.setFont(PDF_FONT_REGULAR, 8.6)
    canvas.setFillColor(colors.HexColor("#24313D"))
    for index, name in enumerate(campaign_summary["Campanha"].head(3).tolist()):
        _draw_fitted_text(canvas, str(name), x + 12, y + height - 38 - index * 14, 216, PDF_FONT_REGULAR, 8.6)
    if report_mode == "awareness":
        metrics = [
            ("Invest. conta", _money(account_totals["spend"])),
            ("Invest. objetivo", _money(focus_totals["spend"])),
            ("Alcance total", _integer(focus_totals["reach"])),
            ("Impressões", _integer(focus_totals["impressions"])),
            ("Cliques", _integer(focus_totals["clicks"])),
            ("CTR", _percent(focus_totals["ctr"])),
            ("CPC médio", _money(focus_totals["cpc"])),
            ("CPM médio", _money(_safe_div(focus_totals["spend"] * 1000, focus_totals["impressions"]))),
            ("Frequência", _ratio(_safe_div(focus_totals["impressions"], focus_totals["reach"]))),
            ("Conversas", _number(focus_totals["messages"])),
        ]
    else:
        metrics = [
            ("Invest. conta", _money(account_totals["spend"])),
            ("Alcance total", _integer(focus_totals["reach"])),
            ("Impressões", _integer(focus_totals["impressions"])),
            ("Total de cliques", _integer(focus_totals["clicks"])),
            ("Conversas", _number(focus_totals["messages"])),
            ("Custo/conv.", _money(focus_totals["cost_per_message"])),
            ("CPC médio", _money(focus_totals["cpc"])),
            ("CTR", _percent(focus_totals["ctr"])),
            ("CPM médio", _money(_safe_div(focus_totals["spend"] * 1000, focus_totals["impressions"]))),
            ("Frequência", _ratio(_safe_div(focus_totals["impressions"], focus_totals["reach"]))),
        ]
    grid_x = x + 238
    cell_w = (width - 252) / 5
    cell_h = 28
    for index, (label, value) in enumerate(metrics):
        col = index % 5
        row = index // 5
        cx = grid_x + col * cell_w
        cy = y + height - 39 - row * cell_h
        canvas.setFillColor(BRAND_SOFT)
        canvas.roundRect(cx, cy, cell_w - 6, 22, 4, fill=1, stroke=0)
        canvas.setFillColor(BRAND_MUTED)
        _draw_fitted_text(canvas, label, cx + 6, cy + 13, cell_w - 18, PDF_FONT_BOLD, 6.7)
        canvas.setFillColor(BRAND_DARK)
        _draw_fitted_text(canvas, value, cx + 6, cy + 4, cell_w - 18, PDF_FONT_BOLD, 8.9)


def _draw_actions_panel(
    canvas,
    x: float,
    y: float,
    width: float,
    height: float,
    account_totals: dict[str, float],
    focus_totals: dict[str, float],
    report_mode: str,
) -> None:
    canvas.setFillColor(colors.white)
    canvas.roundRect(x, y, width, height, 8, fill=1, stroke=0)
    canvas.setFillColor(BRAND_DARK)
    canvas.setFont(PDF_FONT_BOLD, 11.5)
    if report_mode == "awareness":
        canvas.drawString(x + 12, y + height - 18, "Indicadores de reconhecimento")
        rows = [
            ("Investimento da conta", _money(account_totals["spend"]), "Total aplicado"),
            ("Alcance", _integer(focus_totals["reach"]), _money(_safe_div(focus_totals["spend"], focus_totals["reach"]))),
            ("Impressões", _integer(focus_totals["impressions"]), f"{_money(_safe_div(focus_totals['spend'] * 1000, focus_totals['impressions']))} CPM"),
            ("Frequência", _ratio(_safe_div(focus_totals["impressions"], focus_totals["reach"])), "média por pessoa"),
            ("Cliques", _integer(focus_totals["clicks"]), _money(focus_totals["cpc"])),
        ]
    else:
        canvas.drawString(x + 12, y + height - 18, "Conversões e ações por tipo")
        rows = [
            ("Conversas iniciadas", _number(focus_totals["messages"]), _money(focus_totals["cost_per_message"])),
            ("Cliques nos links", _integer(focus_totals["clicks"]), _money(focus_totals["cpc"])),
            ("Pessoas alcançadas", _integer(focus_totals["reach"]), _money(_safe_div(focus_totals["spend"], focus_totals["reach"]))),
            ("Impressões", _integer(focus_totals["impressions"]), f"{_money(_safe_div(focus_totals['spend'] * 1000, focus_totals['impressions']))} CPM"),
        ]
    header = ["Indicador", "Total", "Leitura"] if report_mode == "awareness" else ["Tipo", "Total", "Custo por ação"]
    _draw_compact_table(canvas, x + 12, y + 16, width - 24, height - 44, header, rows, [0.52, 0.22, 0.26])


def _draw_daily_panel(canvas, x: float, y: float, width: float, height: float, daily_summary: pd.DataFrame, report_mode: str) -> None:
    canvas.setFillColor(colors.white)
    canvas.roundRect(x, y, width, height, 8, fill=1, stroke=0)
    canvas.setFillColor(BRAND_DARK)
    canvas.setFont(PDF_FONT_BOLD, 11.5)
    canvas.drawString(
        x + 12,
        y + height - 18,
        "Evolução diária de alcance e impressões" if report_mode == "awareness" else "Evolução diária de conversas e cliques",
    )
    data = daily_summary.tail(8)
    if data.empty:
        canvas.setFillColor(BRAND_MUTED)
        canvas.setFont(PDF_FONT_REGULAR, 8.4)
        canvas.drawString(x + 12, y + height - 44, "Sem dados diários para o período.")
        return
    primary_col = "Alcance" if report_mode == "awareness" else "Mensagens"
    secondary_col = "Impressões" if report_mode == "awareness" else "Cliques"
    max_value = max(float(data[primary_col].max()), float(data[secondary_col].max()), 1)
    chart_x = x + 18
    chart_y = y + 30
    chart_w = width - 36
    group_w = chart_w / len(data)
    for index, (_, row) in enumerate(data.iterrows()):
        base_x = chart_x + index * group_w
        primary_h = (float(row[primary_col]) / max_value) * 52
        secondary_h = (float(row[secondary_col]) / max_value) * 52
        canvas.setFillColor(BRAND_GREEN)
        canvas.roundRect(base_x + 7, chart_y, max(group_w * 0.28, 4), primary_h, 2, fill=1, stroke=0)
        canvas.setFillColor(BRAND_CYAN)
        canvas.roundRect(base_x + 7 + max(group_w * 0.32, 7), chart_y, max(group_w * 0.28, 4), secondary_h, 2, fill=1, stroke=0)
        label = row["Data"].strftime("%d/%m") if hasattr(row["Data"], "strftime") else str(row["Data"])
        canvas.setFillColor(BRAND_MUTED)
        canvas.setFont(PDF_FONT_REGULAR, 6.8)
        canvas.drawCentredString(base_x + group_w / 2, y + 15, label)
    canvas.setFillColor(BRAND_GREEN)
    canvas.rect(x + width - 112, y + height - 24, 7, 7, fill=1, stroke=0)
    canvas.setFillColor(BRAND_MUTED)
    canvas.setFont(PDF_FONT_REGULAR, 7.4)
    canvas.drawString(x + width - 101, y + height - 24, "Alcance" if report_mode == "awareness" else "Conversas")
    canvas.setFillColor(BRAND_CYAN)
    canvas.rect(x + width - 51, y + height - 24, 7, 7, fill=1, stroke=0)
    canvas.setFillColor(BRAND_MUTED)
    canvas.drawString(x + width - 40, y + height - 24, "Impr." if report_mode == "awareness" else "Cliques")


def _draw_featured_table(canvas, x: float, y: float, width: float, height: float, campaign_summary: pd.DataFrame, report_mode: str) -> None:
    canvas.setFillColor(colors.white)
    canvas.roundRect(x, y, width, height, 8, fill=1, stroke=0)
    canvas.setFillColor(BRAND_DARK)
    canvas.setFont(PDF_FONT_BOLD, 11.5)
    canvas.drawString(x + 12, y + height - 18, "Campanhas em destaque")
    rows = []
    for _, row in campaign_summary.iterrows():
        cpm = _safe_div(float(row["Investimento"]) * 1000, float(row["Impressões"]))
        if report_mode == "awareness":
            rows.append(
                (
                    str(row["Campanha"]),
                    str(row["Modelo da Campanha"]),
                    _money(float(row["Investimento"])),
                    _integer(float(row["Alcance"])),
                    _integer(float(row["Impressões"])),
                    _integer(float(row["Cliques"])),
                    _percent(float(row["CTR (%)"])),
                    _money(float(row["CPC"])),
                    _money(cpm),
                    _ratio(float(row["Frequência"])),
                )
            )
        else:
            rows.append(
                (
                    str(row["Campanha"]),
                    str(row["Modelo da Campanha"]),
                    _money(float(row["Custo por Mensagem"])),
                    _money(float(row["Investimento"])),
                    _integer(float(row["Alcance"])),
                    _integer(float(row["Impressões"])),
                    _integer(float(row["Cliques"])),
                    _money(float(row["CPC"])),
                    _money(cpm),
                    _ratio(float(row["Frequência"])),
                )
            )
    if not rows:
        rows = [("Sem campanhas", "-", "-", "-", "-", "-", "-", "-", "-", "-")]
    _draw_compact_table(
        canvas,
        x + 12,
        y + 14,
        width - 24,
        height - 42,
        ["Campanha", "Modelo", "Invest.", "Alcance", "Impressões", "Cliques", "CTR", "CPC", "CPM", "Freq."]
        if report_mode == "awareness"
        else ["Campanha", "Modelo", "Custo/conv.", "Invest.", "Alcance", "Impressões", "Cliques", "CPC", "CPM", "Freq."],
        rows,
        [0.24, 0.16, 0.095, 0.095, 0.10, 0.07, 0.065, 0.065, 0.065, 0.045]
        if report_mode == "awareness"
        else [0.23, 0.16, 0.09, 0.095, 0.09, 0.095, 0.07, 0.07, 0.07, 0.05],
        font_size=6.85,
    )


def _draw_compact_table(
    canvas,
    x: float,
    y: float,
    width: float,
    height: float,
    headers: list[str],
    rows: list[tuple],
    weights: list[float],
    font_size: float = 7.4,
) -> None:
    header_h = 17
    row_h = min(16, (height - header_h) / max(len(rows), 1))
    col_widths = [width * weight / sum(weights) for weight in weights]
    canvas.setFillColor(BRAND_PANEL)
    canvas.roundRect(x, y + height - header_h, width, header_h, 4, fill=1, stroke=0)
    cursor = x
    for index, header in enumerate(headers):
        canvas.setFillColor(colors.white)
        _draw_fitted_text(canvas, header, cursor + 4, y + height - 11.5, col_widths[index] - 8, PDF_FONT_BOLD, font_size)
        cursor += col_widths[index]
    for row_index, row in enumerate(rows):
        row_y = y + height - header_h - (row_index + 1) * row_h
        canvas.setFillColor(colors.white if row_index % 2 == 0 else BRAND_SOFT)
        canvas.rect(x, row_y, width, row_h, fill=1, stroke=0)
        cursor = x
        for col_index, value in enumerate(row):
            canvas.setFillColor(colors.HexColor("#1D2933"))
            if col_index > 1:
                _draw_right_fitted_text(canvas, str(value), cursor + col_widths[col_index] - 4, row_y + 4.5, col_widths[col_index] - 8, PDF_FONT_REGULAR, font_size)
            else:
                _draw_fitted_text(canvas, str(value), cursor + 4, row_y + 4.5, col_widths[col_index] - 8, PDF_FONT_REGULAR, font_size)
            cursor += col_widths[col_index]
        canvas.setStrokeColor(colors.HexColor("#E1E7ED"))
        canvas.line(x, row_y, x + width, row_y)


def _draw_canvas_footer(canvas, margin: float, page_width: float, period: str) -> None:
    canvas.setFillColor(BRAND_MUTED)
    canvas.setFont(PDF_FONT_REGULAR, 7.2)
    canvas.drawString(margin, 12, "Relatório executivo de mídia paga | Gerado automaticamente")
    canvas.drawRightString(page_width - margin, 12, f"Período: {period}")


def _previous_label(value: float, formatter) -> str:
    if not value:
        return "sem período anterior"
    return f"{formatter(value)} no período anterior"


def _draw_fitted_text(canvas, text: str, x: float, y: float, max_width: float, font_name: str, font_size: float) -> None:
    text = _clip_to_width(canvas, text, max_width, font_name, font_size)
    canvas.setFont(font_name, font_size)
    canvas.drawString(x, y, text)


def _draw_right_fitted_text(canvas, text: str, x: float, y: float, max_width: float, font_name: str, font_size: float) -> None:
    text = _clip_to_width(canvas, text, max_width, font_name, font_size)
    canvas.setFont(font_name, font_size)
    canvas.drawRightString(x, y, text)


def _clip_to_width(canvas, text: str, max_width: float, font_name: str, font_size: float) -> str:
    clean = " ".join(str(text or "").split())
    if canvas.stringWidth(clean, font_name, font_size) <= max_width:
        return clean
    suffix = "..."
    while clean and canvas.stringWidth(clean + suffix, font_name, font_size) > max_width:
        clean = clean[:-1]
    return clean.rstrip() + suffix if clean else suffix


def _aggregate(df: pd.DataFrame, group_cols: list[str]) -> pd.DataFrame:
    grouped = df.groupby(group_cols, as_index=False).agg(
        {
            "Impressões": "sum",
            "Alcance": "sum",
            "Cliques": "sum",
            "Investimento": "sum",
            "Mensagens": "sum",
            "Conversões": "sum",
            "Valor de Conversão": "sum",
        }
    )
    grouped["CTR (%)"] = grouped.apply(lambda row: _safe_div(row["Cliques"], row["Impressões"]) * 100, axis=1)
    grouped["Frequência"] = grouped.apply(lambda row: _safe_div(row["Impressões"], row["Alcance"]), axis=1)
    grouped["CPC"] = grouped.apply(lambda row: _safe_div(row["Investimento"], row["Cliques"]), axis=1)
    grouped["Custo por Mensagem"] = grouped.apply(lambda row: _safe_div(row["Investimento"], row["Mensagens"]), axis=1)
    grouped["Custo/Conv."] = grouped.apply(lambda row: _safe_div(row["Investimento"], row["Conversões"]), axis=1)
    grouped["ROAS"] = grouped.apply(lambda row: _safe_div(row["Valor de Conversão"], row["Investimento"]), axis=1)
    return grouped.round({"CTR (%)": 2, "Frequência": 2, "CPC": 2, "Custo por Mensagem": 2, "Custo/Conv.": 2, "ROAS": 2, "Investimento": 2, "Valor de Conversão": 2, "Mensagens": 2, "Conversões": 2})


def _totals(df: pd.DataFrame) -> dict[str, float]:
    spend = float(df["Investimento"].sum())
    conversions = float(df["Conversões"].sum())
    messages = float(df["Mensagens"].sum())
    value = float(df["Valor de Conversão"].sum())
    clicks = float(df["Cliques"].sum())
    impressions = float(df["Impressões"].sum())
    reach = float(df["Alcance"].sum())
    return {
        "spend": spend,
        "reach": reach,
        "impressions": impressions,
        "clicks": clicks,
        "messages": messages,
        "cost_per_message": _safe_div(spend, messages),
        "conversions": conversions,
        "value": value,
        "roas": _safe_div(value, spend),
        "ctr": _safe_div(clicks, impressions) * 100,
        "cpc": _safe_div(spend, clicks),
        "cpa": _safe_div(spend, conversions),
    }


def _totals_from_summary(df: pd.DataFrame) -> dict[str, float]:
    spend = float(df["Investimento"].sum()) if "Investimento" in df else 0.0
    messages = float(df["Mensagens"].sum()) if "Mensagens" in df else 0.0
    clicks = float(df["Cliques"].sum()) if "Cliques" in df else 0.0
    impressions = float(df["Impressões"].sum()) if "Impressões" in df else 0.0
    reach = float(df["Alcance"].sum()) if "Alcance" in df else 0.0
    return {
        "spend": spend,
        "messages": messages,
        "reach": reach,
        "impressions": impressions,
        "clicks": clicks,
        "cpc": _safe_div(spend, clicks),
        "ctr": _safe_div(clicks, impressions) * 100,
        "cost_per_message": _safe_div(spend, messages),
    }


def _pdf_header(styles: dict, period: str) -> list:
    title_table = Table(
        [
            [
                Paragraph("RELATÓRIO EXECUTIVO DE MÍDIA PAGA", styles["HeaderKicker"]),
                Paragraph(f"Período analisado<br/><b>{period}</b>", styles["HeaderMeta"]),
            ],
            [
                Paragraph("Performance de campanhas para mensagens", styles["HeaderTitle"]),
                Paragraph(f"Gerado em<br/><b>{datetime.now().strftime('%d/%m/%Y às %H:%M')}</b>", styles["HeaderMeta"]),
            ],
            [
                Paragraph("Meta Ads e Google Ads | Alcance, impressões, cliques, conversas iniciadas e custo por conversa.", styles["HeaderSubtitle"]),
                "",
            ],
        ],
        colWidths=[183 * mm, 75 * mm],
        hAlign="LEFT",
    )
    title_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, -1), BRAND_DARK),
                ("BOX", (0, 0), (-1, -1), 0.4, colors.HexColor("#24313D")),
                ("LINEBELOW", (0, 0), (-1, 0), 0.5, colors.HexColor("#24313D")),
                ("SPAN", (0, 2), (-1, 2)),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (-1, -1), 14),
                ("RIGHTPADDING", (0, 0), (-1, -1), 14),
                ("TOPPADDING", (0, 0), (-1, -1), 10),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
            ]
        )
    )
    return [title_table, Spacer(1, 12)]


def _comparison_cards(df: pd.DataFrame, previous_df: pd.DataFrame, campaign_summary: pd.DataFrame, styles: dict) -> Table:
    current = _totals(df)
    previous = _totals(previous_df) if not previous_df.empty else {}
    cards = [
        ("Valor investido", _money(current["spend"]), previous.get("spend", 0), current["spend"], _money),
        ("Impressões", _integer(current["impressions"]), previous.get("impressions", 0), current["impressions"], _integer),
        ("Cliques", _integer(current["clicks"]), previous.get("clicks", 0), current["clicks"], _integer),
        ("Número de campanhas", _integer(len(campaign_summary)), len(_aggregate(previous_df, ["Campanha"])) if not previous_df.empty else 0, len(campaign_summary), _integer),
    ]
    cells = []
    for label, value, old_value, new_value, formatter in cards:
        delta = _delta(new_value, old_value)
        previous_label = f"{formatter(old_value)} no período anterior" if old_value else "Sem período anterior"
        cells.append(Paragraph(f"<b>{_escape(label)}</b><br/><font size='15'>{_escape(value)}</font><br/><font color='#35D092'><b>{_escape(delta)}</b></font><br/><font color='#5B6874'>{_escape(previous_label)}</font>", styles["KpiCard"]))
    table = Table([cells], colWidths=[64.5 * mm] * 4, rowHeights=[31 * mm], hAlign="LEFT")
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, -1), colors.white),
                ("BOX", (0, 0), (-1, -1), 0.5, BRAND_LINE),
                ("INNERGRID", (0, 0), (-1, -1), 0.35, BRAND_LINE),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 9),
                ("RIGHTPADDING", (0, 0), (-1, -1), 9),
                ("TOPPADDING", (0, 0), (-1, -1), 8),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
            ]
        )
    )
    return table


def _campaign_overview(campaign_summary: pd.DataFrame, styles: dict) -> Table:
    totals = _totals_from_summary(campaign_summary)
    campaign_names = "<br/>".join(_escape(name) for name in campaign_summary["Campanha"].head(4).tolist()) or "Campanhas selecionadas"
    metrics = [
        ("Alcance total", _integer(totals["reach"])),
        ("Impressões totais", _integer(totals["impressions"])),
        ("Total de cliques", _integer(totals["clicks"])),
        ("Valor investido", _money(totals["spend"])),
        ("CPC médio total", _money(totals["cpc"])),
        ("CTR (todos)", _percent(totals["ctr"])),
        ("CPM médio", _money(_safe_div(totals["spend"] * 1000, totals["impressions"]))),
        ("Frequência", _ratio(_safe_div(totals["impressions"], totals["reach"]))),
    ]
    metric_cells = [
        [Paragraph(label, styles["SmallLabel"]), Paragraph(value, styles["SmallValue"])]
        for label, value in metrics
    ]
    metric_grid = Table(
        [metric_cells[:4], metric_cells[4:]],
        colWidths=[27 * mm] * 4,
        rowHeights=[17 * mm, 17 * mm],
    )
    metric_grid.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, -1), BRAND_SOFT),
                ("INNERGRID", (0, 0), (-1, -1), 0.25, BRAND_LINE),
                ("BOX", (0, 0), (-1, -1), 0.25, BRAND_LINE),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (-1, -1), 6),
                ("RIGHTPADDING", (0, 0), (-1, -1), 6),
            ]
        )
    )
    table = Table(
        [[Paragraph("Campanhas", styles["SectionTitle"]), Paragraph(campaign_names, styles["Body"]), metric_grid]],
        colWidths=[31 * mm, 86 * mm, 136 * mm],
        hAlign="LEFT",
    )
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, -1), colors.white),
                ("BOX", (0, 0), (-1, -1), 0.5, BRAND_LINE),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (-1, -1), 8),
                ("RIGHTPADDING", (0, 0), (-1, -1), 8),
                ("TOPPADDING", (0, 0), (-1, -1), 8),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
            ]
        )
    )
    return table


def _actions_table(df: pd.DataFrame, styles: dict) -> Table:
    totals = _totals(df)
    rows = [
        ["Tipo", "Total", "Custo por ação"],
        ["Conversas iniciadas", _number(totals["messages"]), _money(totals["cost_per_message"])],
        ["Cliques nos links", _integer(totals["clicks"]), _money(totals["cpc"])],
        ["Pessoas alcançadas", _integer(totals["reach"]), _money(_safe_div(totals["spend"], totals["reach"]))],
        ["Impressões", _integer(totals["impressions"]), _money(_safe_div(totals["spend"] * 1000, totals["impressions"])) + " CPM"],
    ]
    return _plain_table(rows, [58 * mm, 32 * mm, 35 * mm], styles)


def _mini_bar_table(df: pd.DataFrame, label_col: str, value_col: str, second_col: str, styles: dict) -> Table:
    rows = [[label_col, value_col, second_col]]
    for _, row in df.iterrows():
        label = row[label_col].strftime("%d/%m") if label_col == "Data" and hasattr(row[label_col], "strftime") else str(row[label_col])
        rows.append([_clip(label, 34), _format_metric(value_col, row[value_col]), _format_metric(second_col, row[second_col])])
    if len(rows) == 1:
        rows.append(["Sem dados", "-", "-"])
    return _plain_table(rows, [58 * mm, 32 * mm, 35 * mm], styles)


def _featured_campaigns_table(df: pd.DataFrame, styles: dict) -> Table:
    rows = [["Campanha", "Modelo", "Custo por conversa", "Valor investido", "Alcance", "Impressões", "Cliques", "CPC", "CPM", "Freq."]]
    for _, row in df.iterrows():
        cpm = _safe_div(float(row["Investimento"]) * 1000, float(row["Impressões"]))
        rows.append(
            [
                _clip(str(row["Campanha"]), 34),
                _clip(str(row.get("Modelo da Campanha", "Não informado")), 26),
                _money(float(row["Custo por Mensagem"])),
                _money(float(row["Investimento"])),
                _integer(float(row["Alcance"])),
                _integer(float(row["Impressões"])),
                _integer(float(row["Cliques"])),
                _money(float(row["CPC"])),
                _money(cpm),
                _ratio(float(row["Frequência"])),
            ]
        )
    table = Table(rows, colWidths=[43 * mm, 34 * mm, 27 * mm, 25 * mm, 23 * mm, 24 * mm, 18 * mm, 17 * mm, 17 * mm, 15 * mm], repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), BRAND_PANEL),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 6.8),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, BRAND_SOFT]),
                ("GRID", (0, 0), (-1, -1), 0.25, BRAND_LINE),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("ALIGN", (1, 1), (-1, -1), "RIGHT"),
                ("LEFTPADDING", (0, 0), (-1, -1), 4),
                ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]
        )
    )
    return table


def _fallback_pdf_story(df: pd.DataFrame, period: str, styles: dict) -> list:
    if df.empty:
        return [
            *_pdf_header(styles, period),
            Paragraph("Nenhum dado encontrado para os filtros selecionados.", styles["Empty"]),
        ]

    campaign_summary = _aggregate(df, ["Plataforma", "Conta", "Modelo da Campanha", "Campanha"]).sort_values(by="Investimento", ascending=False)
    totals = _totals(df)
    story = [
        *_pdf_header(styles, period),
        _simple_kpi_table(totals, styles),
        Spacer(1, 10),
        Paragraph("Resumo consolidado dos resultados", styles["SectionTitle"]),
        _actions_table(df, styles),
        Spacer(1, 10),
        Paragraph("Campanhas em destaque", styles["SectionTitle"]),
        _featured_campaigns_table(campaign_summary.head(10), styles),
        Spacer(1, 8),
        Paragraph(
            "Observação: o relatório usou a versão segura de exportação para preservar o download com todos os principais indicadores.",
            styles["BodyMuted"],
        ),
    ]
    return story


def _simple_kpi_table(totals: dict[str, float], styles: dict) -> Table:
    rows = [
        ["Indicador", "Resultado"],
        ["Investimento total", _money(totals["spend"])],
        ["Conversas iniciadas", _number(totals["messages"])],
        ["Custo por conversa", _money(totals["cost_per_message"])],
        ["Alcance", _integer(totals["reach"])],
        ["Impressões", _integer(totals["impressions"])],
        ["Cliques", _integer(totals.get("clicks", 0))],
        ["CTR", _percent(totals["ctr"])],
        ["CPC", _money(totals["cpc"])],
    ]
    return _plain_table(rows, [78 * mm, 48 * mm], styles)


def _emergency_pdf_bytes(period: str) -> bytes:
    output = BytesIO()
    styles = _pdf_styles()
    document = SimpleDocTemplate(
        output,
        pagesize=landscape(A4),
        rightMargin=18 * mm,
        leftMargin=18 * mm,
        topMargin=14 * mm,
        bottomMargin=14 * mm,
        title="Relatório de Campanhas",
            author="Relatório executivo de mídia paga",
    )
    story = [
        *_pdf_header(styles, period),
        Paragraph("Relatório gerado em modo seguro.", styles["SectionTitle"]),
        Paragraph(
            "Não foi possível montar a versão visual completa do PDF, mas o exportador foi preservado para evitar erro interno. Tente sincronizar novamente os dados e gerar o relatório mais uma vez.",
            styles["BodyMuted"],
        ),
    ]
    document.build(story, onFirstPage=_footer, onLaterPages=_footer)
    return output.getvalue()


def _plain_table(rows: list[list], widths: list[float], styles: dict) -> Table:
    table = Table(rows, colWidths=widths, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), BRAND_PANEL),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 7),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, BRAND_SOFT]),
                ("GRID", (0, 0), (-1, -1), 0.25, BRAND_LINE),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("ALIGN", (1, 1), (-1, -1), "RIGHT"),
                ("LEFTPADDING", (0, 0), (-1, -1), 5),
                ("RIGHTPADDING", (0, 0), (-1, -1), 5),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]
        )
    )
    return table


def _kpi_cards(totals: dict[str, float], styles: dict) -> Table:
    cards = [
        ("Investimento total", _money(totals["spend"]), "Aplicação consolidada no período"),
        ("Conversas", _number(totals["messages"]), "Conversas iniciadas pela campanha"),
        ("Custo por conversa", _money(totals["cost_per_message"]), "Eficiência principal para WhatsApp/Direct"),
        ("CTR / CPC", f"{_percent(totals['ctr'])} | {_money(totals['cpc'])}", "Clique e custo médio por clique"),
    ]
    table = Table(
        [
            [
                [
                    Paragraph(label, styles["KpiLabel"]),
                    Paragraph(value, styles["KpiValue"]),
                    Paragraph(note, styles["KpiNote"]),
                ]
                for label, value, note in cards
            ]
        ],
        colWidths=[64.5 * mm] * 4,
        rowHeights=[26 * mm],
        hAlign="LEFT",
    )
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#F7FAFC")),
                ("BOX", (0, 0), (-1, -1), 0.45, BRAND_LINE),
                ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E5EBF0")),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (-1, -1), 12),
                ("RIGHTPADDING", (0, 0), (-1, -1), 12),
                ("TOPPADDING", (0, 0), (-1, -1), 8),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
            ]
        )
    )
    return table


def _executive_reading(df: pd.DataFrame, campaign_summary: pd.DataFrame, platform_summary: pd.DataFrame, styles: dict) -> list:
    totals = _totals(df)
    best_platform = platform_summary.sort_values(by="Custo por Mensagem", ascending=True).iloc[0]
    top_campaign = campaign_summary.iloc[0]
    bullets = [
        f"No período analisado, o investimento total foi de <b>{_money(totals['spend'])}</b>, gerando <b>{_number(totals['messages'])}</b> conversas iniciadas.",
        f"O custo médio por mensagem ficou em <b>{_money(totals['cost_per_message'])}</b>, com alcance total de <b>{_integer(totals['reach'])}</b> pessoas, <b>{_integer(totals['impressions'])}</b> impressões e frequência média de <b>{_ratio(_safe_div(totals['impressions'], totals['reach']))}</b>.",
        f"A plataforma mais eficiente em custo por mensagem foi <b>{best_platform['Plataforma']}</b>, com <b>{_money(float(best_platform['Custo por Mensagem']))}</b> por conversa iniciada.",
        f"A campanha com maior investimento foi <b>{_escape(str(top_campaign['Campanha']))}</b>, com <b>{_money(float(top_campaign['Investimento']))}</b> aplicados.",
        f"O CTR médio consolidado foi de <b>{_percent(totals['ctr'])}</b> e o CPC médio ficou em <b>{_money(totals['cpc'])}</b>.",
    ]
    return [Paragraph(f"• {item}", styles["Body"]) for item in bullets]


def _data_table(
    df: pd.DataFrame,
    columns: list[str],
    money_cols: set[str] | None = None,
    percent_cols: set[str] | None = None,
    wrap_cols: set[str] | None = None,
    width: float = 258 * mm,
    max_rows: int | None = None,
) -> Table:
    money_cols = money_cols or set()
    percent_cols = percent_cols or set()
    wrap_cols = wrap_cols or set()
    data = df[columns].head(max_rows).copy() if max_rows else df[columns].copy()
    header = [Paragraph(str(col), _pdf_styles()["TableHeader"]) for col in columns]
    rows = []
    for _, item in data.iterrows():
        row = []
        for col in columns:
            value = item[col]
            if col == "Data" and hasattr(value, "strftime"):
                text = value.strftime("%d/%m/%Y")
            elif col in money_cols:
                text = _money(float(value))
            elif col in percent_cols:
                text = _percent(float(value))
            elif col in {"ROAS", "Frequência"}:
                text = _ratio(float(value))
            elif col in {"Alcance", "Impressões", "Cliques"}:
                text = _integer(float(value))
            elif col in {"Mensagens", "Conversões"}:
                text = _number(float(value))
            else:
                text = str(value)
            row.append(Paragraph(_escape(text), _pdf_styles()["TableCell"]) if col in wrap_cols else text)
        rows.append(row)

    col_widths = _column_widths(columns, width)
    table = Table([header] + rows, colWidths=col_widths, repeatRows=1, hAlign="LEFT")
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), BRAND_PANEL),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 7.5),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, BRAND_SOFT]),
                ("GRID", (0, 0), (-1, -1), 0.25, BRAND_LINE),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (-1, -1), 5),
                ("RIGHTPADDING", (0, 0), (-1, -1), 5),
                ("TOPPADDING", (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ]
        )
    )
    for index, col in enumerate(columns):
        if col not in {"Plataforma", "Conta", "Modelo da Campanha", "Campanha", "Data"}:
            table.setStyle(TableStyle([("ALIGN", (index, 1), (index, -1), "RIGHT")]))
    return table


def _column_widths(columns: list[str], width: float) -> list[float]:
    weights = {
        "Campanha": 3.6,
        "Conta": 1.5,
        "Plataforma": 1.25,
        "Data": 1.1,
        "Modelo da Campanha": 1.65,
        "Valor de Conversão": 1.7,
        "Investimento": 1.55,
        "Custo por Mensagem": 1.7,
        "Custo/Conv.": 1.45,
        "Mensagens": 1.2,
        "Impressões": 1.25,
        "Alcance": 1.2,
        "Frequência": 1.2,
    }
    total = sum(weights.get(col, 1.15) for col in columns)
    return [width * (weights.get(col, 1.15) / total) for col in columns]


def _pdf_styles() -> dict:
    base = getSampleStyleSheet()
    base.add(
        ParagraphStyle(
            "HeaderKicker",
            parent=base["Normal"],
            textColor=BRAND_GREEN,
            fontSize=8,
            leading=10,
            fontName="Helvetica-Bold",
            uppercase=True,
        )
    )
    base.add(ParagraphStyle("HeaderTitle", parent=base["Title"], textColor=colors.white, fontSize=24, leading=27, fontName="Helvetica-Bold"))
    base.add(ParagraphStyle("HeaderSubtitle", parent=base["Normal"], textColor=colors.HexColor("#C9D4DD"), fontSize=9.5, leading=13))
    base.add(ParagraphStyle("HeaderMeta", parent=base["Normal"], textColor=colors.HexColor("#DCE6ED"), fontSize=8.5, leading=12, alignment=TA_RIGHT))
    base.add(ParagraphStyle("SectionTitle", parent=base["Heading2"], textColor=BRAND_DARK, fontSize=12, leading=15, spaceBefore=4, spaceAfter=7))
    base.add(ParagraphStyle("Body", parent=base["Normal"], textColor=colors.HexColor("#22303A"), fontSize=9, leading=13, spaceAfter=5))
    base.add(ParagraphStyle("BodyMuted", parent=base["Normal"], textColor=BRAND_MUTED, fontSize=8.4, leading=12, spaceAfter=5))
    base.add(ParagraphStyle("Empty", parent=base["Heading2"], textColor=BRAND_DARK, fontSize=15, leading=18, alignment=TA_CENTER, spaceAfter=8))
    base.add(ParagraphStyle("KpiLabel", parent=base["Normal"], textColor=BRAND_MUTED, fontSize=7.8, leading=9, fontName="Helvetica-Bold"))
    base.add(ParagraphStyle("KpiValue", parent=base["Normal"], textColor=BRAND_DARK, fontSize=13, leading=15, fontName="Helvetica-Bold"))
    base.add(ParagraphStyle("KpiValueLarge", parent=base["Normal"], textColor=BRAND_DARK, fontSize=15.5, leading=18, fontName="Helvetica-Bold"))
    base.add(ParagraphStyle("KpiDelta", parent=base["Normal"], textColor=BRAND_GREEN, fontSize=9, leading=11, fontName="Helvetica-Bold"))
    base.add(ParagraphStyle("KpiNote", parent=base["Normal"], textColor=BRAND_MUTED, fontSize=7.2, leading=9))
    base.add(ParagraphStyle("KpiCard", parent=base["Normal"], textColor=BRAND_DARK, fontSize=8, leading=10.8))
    base.add(ParagraphStyle("SmallLabel", parent=base["Normal"], textColor=BRAND_MUTED, fontSize=6.8, leading=8, fontName="Helvetica-Bold"))
    base.add(ParagraphStyle("SmallValue", parent=base["Normal"], textColor=BRAND_DARK, fontSize=9.4, leading=11, fontName="Helvetica-Bold"))
    base.add(ParagraphStyle("TableHeader", parent=base["Normal"], textColor=colors.white, fontSize=7.2, leading=8.5, fontName="Helvetica-Bold"))
    base.add(ParagraphStyle("TableCell", parent=base["Normal"], textColor=colors.HexColor("#1D2933"), fontSize=7.1, leading=8.5))
    base.add(ParagraphStyle("TableCellRight", parent=base["Normal"], textColor=colors.HexColor("#1D2933"), fontSize=7.1, leading=8.5, alignment=TA_RIGHT))
    return base


def _style_worksheet(worksheet) -> None:
    header_fill = PatternFill("solid", fgColor="111A22")
    header_font = Font(color="FFFFFF", bold=True)
    soft_fill = PatternFill("solid", fgColor="F4F7FA")
    thin = Side(style="thin", color="D9E1E8")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    worksheet.freeze_panes = "A2"
    for row in worksheet.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=False)
            if cell.row == 1:
                cell.fill = header_fill
                cell.font = header_font
            elif cell.row % 2 == 0:
                cell.fill = soft_fill
    for column_cells in worksheet.columns:
        max_length = 0
        column = column_cells[0].column
        for cell in column_cells:
            max_length = max(max_length, len(str(cell.value or "")))
        worksheet.column_dimensions[get_column_letter(column)].width = min(max(max_length + 2, 12), 38)
    worksheet.sheet_view.showGridLines = False


def _footer(canvas, document) -> None:
    canvas.saveState()
    canvas.setStrokeColor(colors.HexColor("#D9E1E8"))
    canvas.line(document.leftMargin, 11 * mm, landscape(A4)[0] - document.rightMargin, 11 * mm)
    canvas.setFillColor(BRAND_MUTED)
    canvas.setFont("Helvetica", 7.5)
    canvas.drawString(document.leftMargin, 6 * mm, "Relatório executivo de mídia paga | Gerado automaticamente")
    canvas.drawRightString(landscape(A4)[0] - document.rightMargin, 6 * mm, f"Página {document.page}")
    canvas.restoreState()


def _platform_label(value: str) -> str:
    return "Google Ads" if value == "google" else "Meta Ads" if value == "meta" else value.title()


def _delta(current: float, previous: float) -> str:
    if not previous:
        return "Novo período"
    change = ((float(current) - float(previous)) / abs(float(previous))) * 100
    sign = "+" if change >= 0 else ""
    return f"{sign}{change:.2f}%".replace(".", ",")


def _format_metric(column: str, value: float) -> str:
    if column in {"Investimento", "CPC", "Custo por Mensagem", "Custo/Conv."}:
        return _money(float(value))
    if column in {"CTR (%)"}:
        return _percent(float(value))
    if column in {"Frequência", "ROAS"}:
        return _ratio(float(value))
    if column in {"Alcance", "Impressões", "Cliques"}:
        return _integer(float(value))
    if column in {"Mensagens", "Conversões"}:
        return _number(float(value))
    return str(value)


def _clip(value: str, max_length: int) -> str:
    clean = " ".join(str(value).split())
    return clean if len(clean) <= max_length else clean[: max_length - 1].rstrip() + "…"


def _objective_label(value: str) -> str:
    normalized = (value or "").strip().upper()
    if normalized in {"OUTCOME_ENGAGEMENT", "MESSAGES", "POST_ENGAGEMENT"}:
        return "Engajamento para mensagens"
    if normalized in {"OUTCOME_LEADS", "LEAD_GENERATION"}:
        return "Geração de cadastros"
    if normalized in {"OUTCOME_TRAFFIC", "LINK_CLICKS"}:
        return "Tráfego"
    if normalized in {"OUTCOME_SALES", "CONVERSIONS"}:
        return "Vendas / conversões"
    if normalized in {"OUTCOME_AWARENESS", "BRAND_AWARENESS", "REACH"}:
        return "Reconhecimento / alcance"
    return value.replace("_", " ").title() if value else "Não informado"


def _period_label(start_date: date | None, end_date: date | None) -> str:
    if start_date and end_date:
        return f"{start_date.strftime('%d/%m/%Y')} a {end_date.strftime('%d/%m/%Y')}"
    return "Período selecionado"


def _safe_div(numerator: float, denominator: float) -> float:
    try:
        if not denominator or math.isclose(float(denominator), 0.0):
            return 0.0
        value = float(numerator) / float(denominator)
        return value if math.isfinite(value) else 0.0
    except (TypeError, ValueError):
        return 0.0


def _money(value: float) -> str:
    return f"R$ {float(value):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


def _integer(value: float) -> str:
    return f"{int(round(float(value))):,}".replace(",", ".")


def _number(value: float) -> str:
    return f"{float(value):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


def _percent(value: float) -> str:
    return f"{float(value):.2f}%".replace(".", ",")


def _ratio(value: float) -> str:
    return f"{float(value):.2f}".replace(".", ",")


def _escape(value: str) -> str:
    return value.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
